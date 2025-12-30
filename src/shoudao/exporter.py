"""
ShouDao CSV exporter - derives CSV from canonical Lead JSON.
"""

import csv
import json
from pathlib import Path
from typing import TextIO
from urllib.parse import urlparse

from .models import Lead, RunResult


def _extract_root_domain(url: str) -> str:
    """Extract root domain from URL (e.g., https://foo.example.com/bar → example.com)."""
    try:
        parsed = urlparse(url)
        netloc = parsed.netloc or parsed.path.split("/")[0]
        # Remove port if present
        netloc = netloc.split(":")[0]
        # Remove www prefix
        if netloc.startswith("www."):
            netloc = netloc[4:]
        return netloc
    except Exception:
        return ""


# CSV column order (stable schema - derived from Lead model)
CSV_COLUMNS = [
    # Organization
    "organization_name",
    "org_type",
    "industries",
    "country",
    "region",
    "city",
    "website",
    "size_indicator",
    "description",
    # Contact (primary)
    "contact_name",
    "contact_title",
    "role_category",
    "email",
    "phone",
    "linkedin",
    "contact_page",
    # Evidence + Quality
    "evidence_urls",
    "evidence_snippets",
    "confidence",
    "dedupe_key",
    "extracted_from_url",
    "domain_aligned",
    "needs_review",
    "buyer_tier",
    "buyer_likelihood",
    # Approach Advice
    "recommended_angle",
    "recommended_first_offer",
    "qualifying_question",
]


def lead_to_row(lead: Lead) -> dict:
    """Convert a Lead to a flat CSV row dict with fallback logic for blank fields."""
    contact = lead.get_primary_contact()

    # Extract channels by type
    email = lead.get_primary_email() or ""
    phone = lead.get_primary_phone() or ""
    linkedin = ""
    contact_page = ""

    if contact:
        for ch in contact.channels:
            if ch.type == "linkedin" and not linkedin:
                linkedin = ch.value
            if ch.type == "contact_page" and not contact_page:
                contact_page = ch.value

    # Collect evidence
    evidence_urls = lead.get_evidence_urls()
    evidence_snippets = []
    for e in lead.evidence:
        if e.snippet:
            evidence_snippets.append(e.snippet)
    for e in lead.organization.evidence:
        if e.snippet:
            evidence_snippets.append(e.snippet)

    # === FALLBACK LOGIC ===
    # Website: if blank, derive from evidence URL or extracted_from_url
    website = str(lead.organization.website) if lead.organization.website else ""
    if not website:
        # Try to derive from extracted_from_url first (most reliable)
        if lead.extracted_from_url:
            website = f"https://{_extract_root_domain(lead.extracted_from_url)}"
        # Otherwise try first evidence URL
        elif evidence_urls:
            website = f"https://{_extract_root_domain(evidence_urls[0])}"

    # Contact page: if blank but we have extracted_from_url, use that
    if not contact_page and lead.extracted_from_url:
        contact_page = lead.extracted_from_url

    # Country: normalize "Unknown" to empty for cleaner CSV
    country = lead.organization.country or ""
    if country.lower() == "unknown":
        country = ""

    return {
        # Organization
        "organization_name": lead.organization.name,
        "org_type": lead.organization.org_type,
        "industries": ";".join(lead.organization.industries),
        "country": country,
        "region": lead.organization.region or "",
        "city": lead.organization.city or "",
        "website": website,
        "size_indicator": lead.organization.size_indicator or "",
        "description": lead.organization.description or "",
        # Contact
        "contact_name": contact.name if contact else "",
        "contact_title": contact.title if contact else "",
        "role_category": contact.role_category if contact else "",
        "email": email,
        "phone": phone,
        "linkedin": linkedin,
        "contact_page": contact_page,
        # Evidence
        "evidence_urls": ";".join(evidence_urls),
        "evidence_snippets": " | ".join(evidence_snippets[:3]),
        "confidence": f"{lead.confidence:.2f}",
        "dedupe_key": lead.dedupe_key or "",
        "extracted_from_url": lead.extracted_from_url or "",
        "domain_aligned": "yes" if lead.domain_aligned else "no",
        "needs_review": "yes" if lead.needs_review else "no",
        "buyer_tier": lead.buyer_tier,
        "buyer_likelihood": f"{lead.buyer_likelihood:.2f}",
        # Advice
        "recommended_angle": lead.advice.recommended_angle if lead.advice else "",
        "recommended_first_offer": lead.advice.recommended_first_offer if lead.advice else "",
        "qualifying_question": lead.advice.qualifying_question if lead.advice else "",
    }


def export_csv(leads: list[Lead], output: Path | TextIO) -> int:
    """Export leads to CSV file. Returns number of rows written."""
    rows = [lead_to_row(lead) for lead in leads]

    if isinstance(output, Path):
        output.parent.mkdir(parents=True, exist_ok=True)
        with open(output, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=CSV_COLUMNS)
            writer.writeheader()
            writer.writerows(rows)
    else:
        writer = csv.DictWriter(output, fieldnames=CSV_COLUMNS)
        writer.writeheader()
        writer.writerows(rows)

    return len(rows)


def export_excel(leads: list[Lead], output: Path) -> int:
    """Export leads to Excel file with auto-fitted column widths."""
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    rows = [lead_to_row(lead) for lead in leads]
    output.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Leads"

    # Write header
    for col_idx, col_name in enumerate(CSV_COLUMNS, 1):
        ws.cell(row=1, column=col_idx, value=col_name)

    # Write data rows
    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, col_name in enumerate(CSV_COLUMNS, 1):
            ws.cell(row=row_idx, column=col_idx, value=row_data.get(col_name, ""))

    # Auto-fit column widths
    for col_idx, col_name in enumerate(CSV_COLUMNS, 1):
        column_letter = get_column_letter(col_idx)

        # Calculate max width from header and data
        max_length = len(col_name)
        for row_data in rows:
            cell_value = str(row_data.get(col_name, ""))
            # Limit cell length consideration to avoid super-wide columns
            cell_length = min(len(cell_value), 50)
            max_length = max(max_length, cell_length)

        # Add a little padding and set width
        adjusted_width = max_length + 2
        ws.column_dimensions[column_letter].width = adjusted_width

    # Freeze header row
    ws.freeze_panes = "A2"

    # Style header row (bold)
    from openpyxl.styles import Font

    for col_idx in range(1, len(CSV_COLUMNS) + 1):
        ws.cell(row=1, column=col_idx).font = Font(bold=True)

    wb.save(output)
    return len(rows)


def export_json(leads: list[Lead], output: Path) -> int:
    """Export leads to JSON file (canonical format)."""
    output.parent.mkdir(parents=True, exist_ok=True)
    data = [lead.model_dump(mode="json") for lead in leads]
    with open(output, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, default=str)
    return len(data)


def generate_report(result: RunResult, output: Path) -> None:
    """Generate a markdown run report."""
    output.parent.mkdir(parents=True, exist_ok=True)

    # Count by country and type
    by_country: dict[str, int] = {}
    by_type: dict[str, int] = {}
    by_industry: dict[str, int] = {}

    for lead in result.leads:
        country = lead.organization.country or "Unknown"
        by_country[country] = by_country.get(country, 0) + 1
        by_type[lead.organization.org_type] = by_type.get(lead.organization.org_type, 0) + 1
        for ind in lead.organization.industries:
            by_industry[ind] = by_industry.get(ind, 0) + 1

    report = f"""# ShouDao Run Report

## Run Info
| Field | Value |
|---|---|
| Run ID | {result.run_id} |
| Started | {result.started_at.isoformat()} |
| Finished | {result.finished_at.isoformat() if result.finished_at else "In progress"} |
| Sources Fetched | {result.sources_fetched} |
| Domains Hit | {result.domains_hit} |
| Total Leads | {len(result.leads)} |

## Prompt
```
{result.config.prompt}
```

## Leads by Country
| Country | Count |
|---|---|
"""
    for country, count in sorted(by_country.items(), key=lambda x: -x[1]):
        report += f"| {country} | {count} |\n"

    report += """
## Leads by Type
| Type | Count |
|---|---|
"""
    for otype, count in sorted(by_type.items(), key=lambda x: -x[1]):
        report += f"| {otype} | {count} |\n"

    if by_industry:
        report += """
## Leads by Industry
| Industry | Count |
|---|---|
"""
        for ind, count in sorted(by_industry.items(), key=lambda x: -x[1]):
            report += f"| {ind} | {count} |\n"

    if result.errors:
        report += "\n## Errors\n"
        for err in result.errors:
            report += f"- {err}\n"

    with open(output, "w", encoding="utf-8") as f:
        f.write(report)
